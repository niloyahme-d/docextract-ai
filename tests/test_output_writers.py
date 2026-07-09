"""Tests for src/output_writers/*.py."""

from __future__ import annotations

import csv

from openpyxl import load_workbook

from src.models import ExtractionMethod, ExtractionResult, ExtractionStatus
from src.output_writers.csv_writer import write_csv
from src.output_writers.excel_writer import write_excel


def _sample_results() -> list[ExtractionResult]:
    return [
        ExtractionResult(
            source_file="invoice_a.pdf",
            method=ExtractionMethod.TEMPLATE,
            status=ExtractionStatus.SUCCESS,
            fields={"vendor_name": "Acme Supplies Ltd", "total_amount": 1366.20},
            confidence=1.0,
            template_id="standard_invoice_v1",
        ),
        ExtractionResult(
            source_file="invoice_b.pdf",
            method=ExtractionMethod.TEMPLATE,
            status=ExtractionStatus.PARTIAL,
            fields={"vendor_name": "Unknown Vendor"},
            confidence=0.3,
            errors=["Required field 'total_amount' not found in document."],
        ),
    ]


def test_write_excel_creates_file_with_expected_rows(config, tmp_path):
    output_path = tmp_path / "results.xlsx"
    results = _sample_results()

    write_excel(results, config, output_path)

    assert output_path.exists()
    wb = load_workbook(str(output_path))
    ws = wb.active

    # Header row + 2 data rows
    assert ws.max_row == 3
    header = [cell.value for cell in ws[1]]
    assert "Vendor Name" in header
    assert "Total Amount" in header


def test_write_excel_flags_non_success_rows(config, tmp_path):
    output_path = tmp_path / "results.xlsx"
    results = _sample_results()

    write_excel(results, config, output_path)

    wb = load_workbook(str(output_path))
    ws = wb.active

    # Row 2 = SUCCESS (no fill), Row 3 = PARTIAL (should be flagged)
    success_row_fill = ws.cell(row=2, column=1).fill.start_color.rgb
    partial_row_fill = ws.cell(row=3, column=1).fill.start_color.rgb
    assert success_row_fill != partial_row_fill


def test_write_csv_creates_file_with_expected_rows(config, tmp_path):
    output_path = tmp_path / "results.csv"
    results = _sample_results()

    write_csv(results, config, output_path)

    assert output_path.exists()
    with output_path.open(newline="", encoding="utf-8") as fh:
        rows = list(csv.reader(fh))

    assert len(rows) == 3  # header + 2 rows
    header = rows[0]
    assert "Vendor Name" in header

    vendor_col = header.index("Vendor Name")
    assert rows[1][vendor_col] == "Acme Supplies Ltd"
    assert rows[2][vendor_col] == "Unknown Vendor"
