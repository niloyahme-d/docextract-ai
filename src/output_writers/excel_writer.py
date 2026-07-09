"""Writes a batch of ExtractionResults to a formatted .xlsx workbook."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.config import AppConfig
from src.models import ExtractionResult, ExtractionStatus

HEADER_FILL = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FLAG_FILL = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")


def write_excel(results: list[ExtractionResult], config: AppConfig, output_path: str | Path) -> Path:
    """Write results to an .xlsx file with one row per document.

    Rows with status != SUCCESS are highlighted so a reviewer can spot
    partial/failed extractions at a glance instead of hunting through logs.
    """
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    field_names = config.all_field_names()
    columns = ["source_file", "status", "method", "confidence", *field_names, "errors"]

    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, result in enumerate(results, start=2):
        row_values = [
            result.source_file,
            result.status.value,
            result.method.value,
            result.confidence,
            *[result.fields.get(name, "") for name in field_names],
            "; ".join(result.errors),
        ]
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if result.status != ExtractionStatus.SUCCESS:
                cell.fill = FLAG_FILL

    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 4)

    ws.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
